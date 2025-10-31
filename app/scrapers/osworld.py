from __future__ import annotations
from typing import Any, Dict, List, Optional
import requests
import io
from datetime import datetime
from openpyxl import load_workbook

OSWORLD_XLSX_URL = "https://os-world.github.io/static/data/osworld_verified_results.xlsx"


def _norm(val: Any) -> Optional[str]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None


def _parse_score(val: Any) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip().replace("%", "")
    try:
        return float(s)
    except Exception:
        return None


def fetch_osworld() -> List[Dict[str, Any]]:
    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
    }
    resp = requests.get(OSWORLD_XLSX_URL, headers=headers, timeout=30)
    resp.raise_for_status()

    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Read header
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows_iter)]

    def find_col(names: list[str]) -> Optional[int]:
        for name in names:
            if name in header:
                return header.index(name)
        return None

    idx_rank = find_col(["rank", "#"])
    idx_model = find_col(["model", "model name"])
    idx_approach = find_col(["approach", "approach type"])
    idx_org = find_col(["org", "organization", "team", "institution"])
    idx_score = find_col(["success rate", "success rate (avg±std)", "accuracy"])
    idx_date = find_col(["date", "submission date"])

    results: List[Dict[str, Any]] = []
    rank_counter = 1

    for row_vals in rows_iter:
        if not any(row_vals):
            continue

        def get(i: Optional[int]) -> Optional[str]:
            if i is None or i >= len(row_vals):
                return None
            return _norm(row_vals[i])

        rank = None
        if idx_rank is not None:
            try:
                rank = int(get(idx_rank)) if get(idx_rank) else rank_counter
            except Exception:
                rank = rank_counter
        else:
            rank = rank_counter
        rank_counter += 1

        model = get(idx_model)
        approach = get(idx_approach)
        org = get(idx_org)
        date = get(idx_date)

        score_val = None
        if idx_score is not None:
            raw_score = get(idx_score)
            if raw_score:
                # format: "25.6±2.3" or "25.6"
                score_val = _parse_score(raw_score.split("±")[0])

        # Convert datetime objects in row_vals to strings for JSON serialization
        row_safe = []
        for v in row_vals:
            if isinstance(v, datetime):
                row_safe.append(v.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                row_safe.append(v)

        results.append({
            "bench": "osworld",
            "rank": rank,
            "agent": approach,
            "model": model,
            "date": date,
            "agent_org": None,
            "model_org": None,
            "org": org,
            "score": score_val,
            "score_error": None,
            "raw": {"row": row_safe, "header": header},
        })

    wb.close()
    return results
